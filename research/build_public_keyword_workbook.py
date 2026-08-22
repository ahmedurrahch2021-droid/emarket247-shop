"""Build a public-data keyword workbook with qualitative evidence instead of fabricated search metrics."""
import csv
from collections import Counter
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
INPUT = ROOT / "BANGLADESH_PUBLIC_KEYWORD_MATRIX.csv"
OUTPUT = ROOT / "eMarket247_Bangladesh_Public_Keyword_Research.xlsx"

with INPUT.open(encoding="utf-8", newline="") as handle:
    keywords = list(csv.DictReader(handle))

assert len(keywords) == 300, f"Expected 300 keyword rows, received {len(keywords)}"

navy = "1F2937"
red = "ED1C24"
warm = "F6F2EE"
white = "FFFFFF"
border = Border(left=Side(style="thin", color="DED7D0"), right=Side(style="thin", color="DED7D0"), top=Side(style="thin", color="DED7D0"), bottom=Side(style="thin", color="DED7D0"))

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

def style_data(ws, start_row, end_row, cols):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=cols):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

# Tab 1: Topic Cluster Summary
summary = Workbook()
ws = summary.active
ws.title = "Topic Cluster Summary"
ws.merge_cells("A1:H1")
ws["A1"] = "eMarket247 — Bangladesh Public Keyword Research"
ws["A1"].font = Font(size=15, bold=True, color=navy)
ws.merge_cells("A2:H2")
ws["A2"] = "Scope: Bangladesh | Languages: Bengali and English | Data: public qualitative research, not measured keyword-volume data"
ws["A2"].font = Font(italic=True, color="5C534D")
ws.merge_cells("A3:H3")
ws["A3"] = "Important: N/A means no verified public metric was available. No estimated volume, CPC, difficulty, traffic share, or domain rating is presented as measured data."
ws["A3"].font = Font(italic=True, color=red)
headers = ["Topic Cluster", "Keywords", "High Priority", "Medium Priority", "English", "Bengali", "Measured Volume", "Implementation Purpose"]
for idx, header in enumerate(headers, 1):
    ws.cell(row=5, column=idx, value=header)
style_header(ws, 5, len(headers))
cluster_order = [
    "Product category discovery", "Occasion and seasonal discovery", "Trust, policy, and purchase readiness", "Style, look, and material exploration", "Guides, FAQs, and answer-engine content",
]
purposes = {
    "Product category discovery": "Individual bilingual category pages and category-menu subpages.",
    "Occasion and seasonal discovery": "Puja, bridal, gifting, and occasion-led collection pages.",
    "Trust, policy, and purchase readiness": "Care, policy, product-detail, support, and purchase-readiness content.",
    "Style, look, and material exploration": "Editorial discovery pages and category-filter language.",
    "Guides, FAQs, and answer-engine content": "Bilingual guides, FAQs, and concise answer-first content.",
}
for row_index, cluster in enumerate(cluster_order, 6):
    scoped = [item for item in keywords if item["cluster"] == cluster]
    ws.append([cluster, len(scoped), sum(item["priority"] == "High" for item in scoped), sum(item["priority"] == "Medium" for item in scoped), sum(item["language"] == "English" for item in scoped), sum(item["language"] == "Bengali" for item in scoped), "N/A — public data", purposes[cluster]])
style_data(ws, 6, 10, len(headers))
ws.append(["TOTAL", len(keywords), sum(item["priority"] == "High" for item in keywords), sum(item["priority"] == "Medium" for item in keywords), sum(item["language"] == "English" for item in keywords), sum(item["language"] == "Bengali" for item in keywords), "N/A — public data", "Bilingual architecture and content roadmap"])
for cell in ws[11]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="FDE8E8")
    cell.border = border
widths = [40, 12, 14, 16, 12, 12, 24, 48]
for index, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(index)].width = width

# Tab 2: 300 Keyword targets
target = summary.create_sheet("Priority Keyword Targets")
target.append(["Keyword", "Topic Cluster", "Language", "Search Intent", "Measured Volume (Bangladesh)", "KD", "CPC", "Target Page", "Priority", "Public Evidence", "Metric Note"])
style_header(target, 1, 11)
for item in keywords:
    target.append([item["keyword"], item["cluster"], item["language"], item["intent"], "N/A — public data", "N/A — public data", "N/A — public data", item["page"], item["priority"], item["evidence"], item["metric_note"]])
style_data(target, 2, 301, 11)
for row in range(2, 302):
    priority = target.cell(row=row, column=9)
    priority.fill = PatternFill("solid", fgColor="DFF3E3" if priority.value == "High" else "FFF3CD")
target.freeze_panes = "A2"
target.auto_filter.ref = "A1:K301"
for index, width in enumerate([42, 38, 12, 18, 24, 18, 18, 32, 12, 38, 44], 1):
    target.column_dimensions[get_column_letter(index)].width = width

# Tab 3: qualitative landscape and gaps
landscape = summary.create_sheet("Competitor Gaps & Landscape")
landscape["A1"] = "Public qualitative market-reference landscape"
landscape["A1"].font = Font(size=14, bold=True, color=navy)
landscape.merge_cells("A1:E1")
landscape["A2"] = "These are publicly visible market references, not a measured top-competitor or traffic-ranking list. Traffic share, traffic value, and authority metrics are intentionally marked N/A."
landscape.merge_cells("A2:E2")
landscape["A2"].alignment = Alignment(wrap_text=True)
landscape["A2"].font = Font(italic=True, color=red)
landscape.append([])
landscape.append(["Public market reference", "Observed focus", "Traffic share", "Traffic value", "Observation"])
style_header(landscape, 4, 5)
references = [
    ("aarong.com", "Jewellery category hierarchy", "N/A", "N/A", "Public category taxonomy includes earrings, necklaces, sets, bracelets/bangles, rings, anklets, lockets, and pendants."),
    ("kunjojewellers.com", "Gold and diamond categories", "N/A", "N/A", "Public category/store/FAQ/account/cookie pattern."),
    ("pearlartistry.com", "Fashion/pearl product detail", "N/A", "N/A", "Nested category, product, compare, wishlist, and detail patterns."),
    ("gauravjewellers.com", "Gold jewellery and trust prompts", "N/A", "N/A", "Public certification, shipping, exchange, payment, and trade-license messaging patterns."),
    ("diamondworldltd.com", "Gold and diamond jewellery", "N/A", "N/A", "Public premium-jewellery search positioning."),
    ("alaminjewellers.com", "Gold jewellery", "N/A", "N/A", "Bilingual category and current-price context."),
    ("chowdhurygold.com", "Online gold jewellery", "N/A", "N/A", "Public category and delivery messaging context."),
    ("othoba.com", "Marketplace jewellery", "N/A", "N/A", "Broad category/page and price-led marketplace context."),
    ("daraz.com.bd", "Marketplace jewellery", "N/A", "N/A", "Bengali category query language and marketplace behavior."),
    ("simplicy.com.bd", "Fashion accessories", "N/A", "N/A", "Category taxonomy for accessories, sets, and gifts."),
    ("moonflowerbd.com", "Minimal fashion jewellery", "N/A", "N/A", "Minimal-jewellery and women's fashion-accessories positioning."),
    ("utshob.com", "Gifts and jewellery", "N/A", "N/A", "Gift-context discovery reference."),
]
for ref in references:
    landscape.append(ref)
style_data(landscape, 5, 4 + len(references), 5)
gap_start = 7 + len(references)
landscape.cell(row=gap_start, column=1, value="Qualitative content-gap opportunities").font = Font(size=12, bold=True)
landscape.merge_cells(start_row=gap_start, start_column=1, end_row=gap_start, end_column=5)
gap_headers = ["Opportunity", "Language", "Intent", "Measured Volume", "Recommended action"]
for index, header in enumerate(gap_headers, 1):
    landscape.cell(row=gap_start + 2, column=index, value=header)
style_header(landscape, gap_start + 2, 5)
gaps = [
    ("Puja jewellery collection with clear bilingual discovery", "Bengali + English", "Commercial", "N/A — public data", "Create reciprocal Bengali/English pre-Puja landing pages in September, then update with approved collection records."),
    ("Bangles versus bracelets explainer", "Bengali + English", "Informational", "N/A — public data", "Create answer-first guide and internal links to both category pages."),
    ("Ring-size preparation guide", "Bengali + English", "Informational", "N/A — public data", "Publish only approved measurement method and product-specific fit information."),
    ("Earring style guide for sari, Puja, and everyday looks", "Bengali + English", "Informational", "N/A — public data", "Create visible, editorial recommendation page without fabricated product claims."),
    ("Jewellery gifting guide", "Bengali + English", "Commercial", "N/A — public data", "Create occasion, recipient, and category discovery paths."),
    ("Product-image transparency and detail standard", "Bengali + English", "Trust", "N/A — public data", "State the publishing standard only after it matches operating practice."),
    ("Care and materials guidance", "Bengali + English", "Informational", "N/A — public data", "Publish material-specific content only after approved material data exists."),
    ("Secure purchase and support readiness", "Bengali + English", "Trust", "N/A — public data", "Use clear support and policy pages; do not claim unapproved payment or delivery services."),
    ("Necklace-length and layering guide", "Bengali + English", "Informational", "N/A — public data", "Create editorial guide linked to necklaces and sets."),
    ("Bridal jewellery discovery", "Bengali + English", "Commercial", "N/A — public data", "Create category-led bridal edit without price or stock claims."),
    ("Gift jewellery discovery", "Bengali + English", "Commercial", "N/A — public data", "Create recipient/occasion navigation and request-for-detail pathway."),
    ("Bengali category aliases in navigation and internal links", "Bengali", "Navigational", "N/A — public data", "Use visible Bengali labels while keeping English canonical URLs."),
    ("Puja gifting timeline", "Bengali + English", "Informational", "N/A — public data", "Explain seasonal discovery without delivery-date promises."),
    ("Jewellery category comparison hub", "Bengali + English", "Informational", "N/A — public data", "Build concise comparison content with linked category pages."),
    ("Occasion-led jewellery hub", "Bengali + English", "Commercial", "N/A — public data", "Use wedding, Puja, birthday, anniversary, and gifting routes."),
    ("Why product detail matters", "Bengali + English", "Trust", "N/A — public data", "Explain image, size, material, and care detail standards when those fields are verified."),
    ("Bilingual FAQ hub", "Bengali + English", "Informational", "N/A — public data", "Use visible questions and concise answers, not unsupported FAQ-rich-result promises."),
    ("In-store / online support page", "Bengali + English", "Navigational", "N/A — public data", "Publish only verified contact, address, and support hours."),
    ("Product-review policy", "Bengali + English", "Trust", "N/A — public data", "Introduce only when genuine review collection and moderation exist."),
    ("Seasonal collection archive", "Bengali + English", "Informational", "N/A — public data", "Keep evergreen explanatory content and replace temporary offers with current approved information."),
]
for gap in gaps:
    landscape.append(gap)
style_data(landscape, gap_start + 3, gap_start + 2 + len(gaps), 5)
for index, width in enumerate([39, 22, 18, 22, 67], 1):
    landscape.column_dimensions[get_column_letter(index)].width = width

summary.save(OUTPUT)
print(f"Created {OUTPUT}")
